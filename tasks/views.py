

import os
from django.shortcuts import render, redirect
from django.http import HttpResponse, FileResponse
# cuando ejecuta devuelve un formulario
from django.contrib.auth.forms import UserCreationForm, AuthenticationForm
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from .models import *
from .forms import *
from django.views.generic.detail import DetailView


from openpyxl import Workbook
from openpyxl.utils import get_column_letter






# Create your views here.
"""
def helloworld(request):
    title = 'helloworld'
    return render(request, 'index.html', {'mytitle': title
    })
Vemos que dentro de esto, se va renderizar un html, y tenemos una variable,
llamada title, que dentro del html podemos llamarla como 'mytitle'.
"""
## BASE HTML CONTENT ##

## HTMLS TEMPLATES ## 

def index(request):
    novedades = Novedad.objects.order_by('-fecha')
    return render(request, 'index.html', {'novedades': novedades})
            
###################### LOGS HTML #####################
def signup(request):
    # aca indicamos, si entra por POST realiza el envio de datos para la bd
    if request.method == 'POST':
        # Si password1 es igual a password2 entonces intentamos guardar
        if request.POST['password1'] == request.POST['password2']:
            # Hacemos un try except para que la app no rompa todo la pagina en el caso de error, porque al generar dos ->
            # Usuarios  iguales podes tener error 'unique' que se genera cuando creas dos usuarios iguales y no es posible eso en la db
            try:
                user = User.objects.create_user(
                    username=request.POST['username'], password=request.POST['password1'])
                # Crea   la cokie de login, para que el navegador sepa que estas logeado
                user.save()
                login(request, user)
                return redirect('index')  # Redirect to home page

            except:
                return render(request, 'signup.html', {
                    'form': UserCreationForm,
                    'error': 'Failed to signup username'
                })

        return render(request, 'signup.html', {
            'form': UserCreationForm,
            'error': 'password does not match'
        })

    else:  # Aca indicamos que si entra por GET nos haga el render de form
        return render(request, 'signup.html', {
            'form': UserCreationForm
        })

def signout(request):
    logout(request)
    return redirect('index')

def signin(request):
    # Si entramos por GET, mandamos a la persona al form
    if request.method == 'GET':
        return render(request, 'signin.html', {
            'form': AuthenticationForm
        })

    else:
        # Si entramos por POST, tomamos los datos del form, usuario+pasword e igualamos, vemos que sean correctos.
        user = authenticate(request, username=request.POST.get(
            'username'), password=request.POST.get('password'))
        if user is None:
            # Si no son correctos lo mandamos a la vista de logeo.
            return render(request, 'signin.html', {
                'form': AuthenticationForm,
                'error': 'Invalid username or password'
            })

        else:
            # Si son correctos, guarda su sesion, en cokies, y direccionalo al home.
            login(request, user)
            return redirect('index')
################### FUNCIONALIDADES ####################

@login_required
def area_soporte(request):
    area_soporte = Referentes_soporte.objects.filter(nombre_referente="Adrian Martinez")
    return render ( request, "area_soporte.html",{"area_soporte": area_soporte}) 

@login_required
def area_desarrollo(request):
    area_desarrollo = Referentes_desarrollo.objects.filter(nombre_referente="Dario Waicen")
    return render ( request, "area_desarrollo.html",{"area_desarrollo": area_desarrollo}) 

@login_required
def area_swbase(request):
    area_swbase = Referentes_swbase.objects.filter(nombre_referente="Leonardo Raurell")
    return render ( request, "area_swbase.html",{"area_swbase": area_swbase}) 

@login_required
def area_redes(request):
    area_redes = Referentes_redes.objects.filter(nombre_referente="Bartolomeo Gustavo")
    return render ( request, "area_redes.html",{"area_redes": area_redes}) 

@login_required
def area_tel(request):
    area_tel = Referentes_tel.objects.filter(nombre_referente="Bartolomeo Gustavo")
    return render ( request, "area_tel.html",{"area_tel": area_tel}) 

@login_required
def area_seginf(request):
    area_seginf = Referentes_seginf.objects.filter(nombre_referente="Carbone Esteban")
    return render ( request, "area_seginf.html",{"area_seginf": area_seginf}) 

@login_required
def formularios(request):
    formularios = Formularios.objects.filter(nombre_referente="Mariana Ferreyra")
    return render ( request, "formularios.html",{"formularios": formularios}) 

@login_required
def capacitacion(request):
    capacitacion = Capacitacion.objects.filter(nombre_referente="Matias Alegre")
    return render ( request, "capacitacion.html",{"capacitacion": capacitacion}) 
@login_required
def flow(request):
    return render(request, 'flow.html')

def incidentes(request):
    return render(request, 'incidentes.html')


def referentes(request):
    return render(request, 'referentes.html')


class ver_formularios(LoginRequiredMixin, DetailView): 
    model = Formularios
    template_name = "ver_formularios.html"


class ver_errores_soporte(LoginRequiredMixin, DetailView): 
    model = Referentes_soporte
    template_name = "ver_errores_soporte.html"


class ver_errores_desarrollo(LoginRequiredMixin, DetailView): 
    model = Referentes_desarrollo
    template_name = "ver_errores_desarrollo.html"

 
class ver_errores_swbase(LoginRequiredMixin, DetailView): 
    model = Referentes_swbase
    template_name = "ver_errores_swbase.html"


class ver_errores_redes(LoginRequiredMixin, DetailView): 
    model = Referentes_redes
    template_name = "ver_errores_redes.html"

class ver_errores_tel(LoginRequiredMixin, DetailView): 
    model = Referentes_tel
    template_name = "ver_errores_tel.html"
    

class ver_errores_seginf(LoginRequiredMixin, DetailView): 
    model = Referentes_seginf
    template_name = "ver_errores_seginf.html"
    

class ver_capacitacion(LoginRequiredMixin, DetailView): 
    model = Capacitacion
    template_name = "ver_capacitacion.html"
    
##aplicacion para EXCEL form

def formulario(request):
    return render(request, 'formulario.html')

def error_excel(request):
    return render(request, 'error_excel.html')

def exito(request):
    return render(request, 'exito.html')


def generar_excel(request):
    # Obtener los datos del formulario
    titulo = request.POST.get('titulo')
    reclamante = request.POST.get('reclamante')
    afectado = request.POST.get('afectado')
    contacto = request.POST.get('contacto')

    # Crear el libro de trabajo de Excel
    wb = Workbook()
    ws = wb.active

    # Agregar los datos al archivo Excel
    datos = [('Título de error', titulo),
             ('Nombre de reclamante', reclamante),
             ('Nombre de afectado', afectado),
             ('Número de contacto', contacto)]

    for row in range(len(datos)):
        label, value = datos[row]
        col_label = get_column_letter(1)
        col_value = get_column_letter(2)

        cell_label = '{}{}'.format(col_label, row+1)
        cell_value = '{}{}'.format(col_value, row+1)

        ws[cell_label] = label
        ws[cell_value] = value

    # Guardar el archivo Excel
    ruta_archivos = os.getenv('RUTA_FORM_RECLAMOS')
    file_path = os.path.join(ruta_archivos, 'nuevo.xlsx')
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    wb.save(file_path)

    return render(request, 'exito.html')

######## PANEL DE BUSQUEDA #############

def resultados_de_busqueda(request):
    query = request.GET.get('query', '')

    # Realizar la búsqueda en varios modelos
    resultados_modelo1 = Referentes_soporte.objects.filter(titulo_error__icontains=query)
    resultados_modelo2 = Referentes_desarrollo.objects.filter(titulo_error__icontains=query)
    resultados_modelo3 = Referentes_swbase.objects.filter(titulo_error__icontains=query)
    resultados_modelo4 = Referentes_redes.objects.filter(titulo_error__icontains=query)
    resultados_modelo5 = Referentes_tel.objects.filter(titulo_error__icontains=query)
    resultados_modelo6 = Referentes_seginf.objects.filter(titulo_error__icontains=query)
    resultados_modelo7 = Formularios.objects.filter(titulo_error__icontains=query)

    # Combinar los resultados en una lista de diccionarios
    resultados = []
    for resultado in resultados_modelo1:
        resultados.append({'modelo': 'Modelo1', 'objeto': resultado})
        
    for resultado in resultados_modelo2:
        resultados.append({'modelo': 'Modelo2', 'objeto': resultado})
    
    for resultado in resultados_modelo3:
        resultados.append({'modelo': 'Modelo3', 'objeto': resultado})
    
    for resultado in resultados_modelo4:
        resultados.append({'modelo': 'Modelo4', 'objeto': resultado})
    
    for resultado in resultados_modelo5:
        resultados.append({'modelo': 'Modelo5', 'objeto': resultado})
    
    for resultado in resultados_modelo6:
        resultados.append({'modelo': 'Modelo6', 'objeto': resultado})
    
    for resultado in resultados_modelo7:
        resultados.append({'modelo': 'Modelo7', 'objeto': resultado})
    


    return render(request, 'search.html', {'resultados': resultados})



######### novedades 
